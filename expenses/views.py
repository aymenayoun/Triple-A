from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from userpreferences.models import UserPreference
from .models import Category,Expenses
from django.core.paginator import Paginator
import json
from django.http import HttpResponse, JsonResponse
from django.utils.dateparse import parse_date
import datetime
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from openpyxl.drawing.image import Image
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile
from django.db.models import Sum
from django.templatetags.static import static
from django.utils.http import urlsafe_base64_encode
from django.conf import settings
# Create your views here.

@login_required(login_url='/authentication/login')
def index(request):
    categories=Category.objects.all()
    expenses=Expenses.objects.filter(owner=request.user)
    paginator=Paginator(expenses,5)#paginator:object/Paginator:class/expenses:to show in pages/2:num of items per page
    page_number=request.GET.get('page')#pick the value of page from the url
    page_obj=Paginator.get_page(paginator,page_number)#the page to view
    currency=UserPreference.objects.get(user=request.user).currency
    currencystr=str(currency)
    currencyabr=currencystr[:3]
    context={'expenses': expenses,'page_obj':page_obj,'currency':currency,'currencyabr':currencyabr}#to have access to context in the html
    return render(request,'expenses/index.html',context)

@login_required(login_url='/authentication/login')
def add_expenses(request):
    categories = Category.objects.all()
    context = {'categories': categories}

    if request.method == 'POST':
        amount = request.POST.get('amount')
        date = request.POST.get('expense_date')  
        description = request.POST.get('description', '')
        category = request.POST.get('category')
        if not amount:
            messages.error(request, 'Please specify the amount.')
        elif not date:
            messages.error(request, 'Please specify the date.')
        elif not category:
            messages.error(request, 'Please select a category.')
        else:
            Expenses.objects.create(
                amount=amount,
                date=date,
                description=description,  # This will be an empty str if not provided
                owner=request.user,
                category=category
            )
            messages.success(request, 'Expense added successfully!')

            return redirect('expenses')
        return render(request, 'expenses/add_expenses.html', context)
    return render(request, 'expenses/add_expenses.html', context)
    
    

@login_required(login_url='/authentication/login')
def edit_expense(request, id):
    expense = get_object_or_404(Expenses, pk=id, owner=request.user)
    categories = Category.objects.all()

    if request.method == 'POST':
        amount = request.POST.get('amount')
        date = request.POST.get('expense_date')
        description = request.POST.get('description')
        category = request.POST.get('category')

        if not amount:
            messages.error(request, 'Please specify the amount.')
        elif not date:
            messages.error(request, 'Please specify the date.')
        elif not category:
            messages.error(request, 'Please select a category.')
        else:
            expense.amount = amount
            expense.date = date
            expense.description = description
            expense.category = category
            expense.save()
            messages.success(request, 'Expense updated successfully!')
            return redirect('expenses')  # Redirect to the list of expenses

    context = {
        'expense': expense,
        'categories': categories,
        'values': expense
    }
    return render(request, 'expenses/edit_expense.html', context)



@login_required(login_url='/authentication/login')
def delete_expense(request, id):
    expense = get_object_or_404(Expenses, pk=id, owner=request.user)
    expense.delete()
    messages.success(request, 'Expense deleted successfully!')
    return redirect('expenses')



@login_required(login_url='/authentication/login')
def search_expenses(request):
    if request.method == 'POST':
        search_str = json.loads(request.body).get('searchText')
        

        try:
            date = parse_date(search_str)
        except ValueError:
            date = None

        if date:
            expenses = Expenses.objects.filter(date=date, owner=request.user)
        else:
            
            expenses = Expenses.objects.filter(
                amount__startswith=search_str, owner=request.user) | Expenses.objects.filter(
                description__icontains=search_str, owner=request.user) | Expenses.objects.filter(
                category__icontains=search_str, owner=request.user)

        data = expenses.values()
        return JsonResponse(list(data), safe=False)
    
@login_required(login_url='/authentication/login')    
def expense_category_summary(request):
    today_date = datetime.date.today()
    six = today_date - datetime.timedelta(days=180)
    expenses = Expenses.objects.filter(owner=request.user, date__gte=six, date__lte=today_date)

    finalrep = {}
    
    def get_category(expense):
        return expense.category

    def get_expense_category_amount(category):
        amount = 0
        filtered_by_category = expenses.filter(category=category)
        for exp in filtered_by_category:
            amount += exp.amount
        return amount

    category_list = list(set(map(get_category, expenses)))
    for category in category_list:
        finalrep[category] = get_expense_category_amount(category)

    return JsonResponse({'expense_category_data': finalrep}, safe=False)


@login_required(login_url='/authentication/login')
def stats_view(request):
    return render(request,'expenses/expstats.html')
        
    
@login_required(login_url='/authentication/login')    
def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    date=str(datetime.datetime.now())
    response['Content-Disposition'] = 'attachment; filename=expenses'+date+'.csv'
    writer=csv.writer(response)
    writer.writerow(['Amount','Category','Description','Date'])
    expenses=Expenses.objects.filter(owner=request.user)
    for exp in expenses:
        writer.writerow([exp.amount,exp.category,exp.description,exp.date])
   
    return response

@login_required(login_url='/authentication/login')
def export_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date = str(datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
    response['Content-Disposition'] = f'attachment; filename=expenses_{date}.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Expenses'
    ws['A1'] = 'Amount'
    ws['B1'] = 'Category'
    ws['C1'] = 'Description'
    ws['D1'] = 'Date'
    expenses=Expenses.objects.filter(owner=request.user)
    date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD')
    for i, exp in enumerate(expenses, start=2):
        ws.cell(row=i, column=1).value = exp.amount
        ws.cell(row=i, column=2).value = exp.category
        ws.cell(row=i, column=3).value = exp.description
        
        date_cell = ws.cell(row=i, column=4)
        if isinstance(exp.date, datetime.date): 
            date_cell.value = exp.date
        else:
            
            date_cell.value = datetime.datetime.strptime(str(exp.date), "%Y-%m-%d")
        date_cell.number_format = 'YYYY-MM-DD'
        for col in range(1, 5): 
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        ws.column_dimensions[column].width = max_length + 2  
    wb.save(response)
    return response

@login_required(login_url='/authentication/login')
def export_pdf(request):
    expenses = Expenses.objects.filter(owner=request.user)
    date = str(datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=expenses_{date}.pdf'
    response['Content-Transfer-Encoding'] = "binary"
    expenses=Expenses.objects.filter(owner=request.user)
    logo_url = request.build_absolute_uri(static('img/Triple_A_icon.jpg'))
    sum=expenses.aggregate(total=Sum('amount'))['total'] or 0
    html_string = render_to_string('expenses/pdf-output.html', {
        'expenses': expenses,
        'total': sum,
        'logo_url': logo_url
    })

    
    html = HTML(string=html_string)

    with tempfile.NamedTemporaryFile(delete=False) as output:
        html.write_pdf(output)
        output.seek(0)
        response.write(output.read())

    output.close()
    return response