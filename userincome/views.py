from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from userpreferences.models import UserPreference
from .models import Source, UserIncome
from django.core.paginator import Paginator
import json
from django.http import HttpResponse, JsonResponse
from django.utils.dateparse import parse_date
import datetime
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from openpyxl.drawing.image import Image
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile
from django.db.models import Sum
from django.templatetags.static import static
from django.utils.http import urlsafe_base64_encode
from django.conf import settings
@login_required(login_url='/authentication/login')
def index(request):
    sources = Source.objects.all()
    income = UserIncome.objects.filter(owner=request.user)
    paginator = Paginator(income, 5)
    page_number = request.GET.get('page')
    page_obj = Paginator.get_page(paginator, page_number)
    currency = UserPreference.objects.get(user=request.user).currency
    currencystr = str(currency)
    currencyabr = currencystr[:3]
    context = {'income': income, 'page_obj': page_obj, 'currency': currency, 'currencyabr': currencyabr}
    return render(request, 'income/index.html', context)

@login_required(login_url='/authentication/login')
def add_income(request):
    sources = Source.objects.all()  # Fetch all available sources
    context = {'sources': sources, 'values': request.POST}

    if request.method == 'POST':
        amount = request.POST.get('amount')
        date = request.POST.get('income_date')  
        description = request.POST.get('description')
        source_name = request.POST.get('source')  

        if not amount:
            messages.error(request, 'Please specify the amount.')
        elif not date:
            messages.error(request, 'Please specify the date.')
        elif not source_name:
            messages.error(request, 'Please select a source.')
        else:
            
            UserIncome.objects.create(
                amount=amount,
                date=date,
                description=description,
                owner=request.user,
                source=source_name  
            )
            messages.success(request, 'Income added successfully!')
            return redirect('income')

        return render(request, 'income/add_income.html', context)
    
    return render(request, 'income/add_income.html', context)


@login_required(login_url='/authentication/login')
def edit_income(request, id):
    income = get_object_or_404(UserIncome, pk=id, owner=request.user)
    sources = Source.objects.all()

    if request.method == 'POST':
        amount = request.POST.get('amount')
        date = request.POST.get('income_date')
        description = request.POST.get('description', '')
        source_name = request.POST.get('source')

        if not amount:
            messages.error(request, 'Please specify the amount.')
        elif not date:
            messages.error(request, 'Please specify the date.')
        elif not source_name:
            messages.error(request, 'Please select a source.')
        else:
            income.source = source_name
            income.amount = amount
            income.date = date
            income.description = description
            income.save()
            messages.success(request, 'Income updated successfully!')
            return redirect('income')

    context = {
        'income': income,
        'sources': sources,
        'values': income
    }
    return render(request, 'income/edit_income.html', context)

@login_required(login_url='/authentication/login')
def delete_income(request, id):
    income = get_object_or_404(UserIncome, pk=id, owner=request.user)
    income.delete()
    messages.success(request, 'Income record removed.')
    return redirect('income')


@login_required(login_url='/authentication/login')
def search_income(request):
    if request.method == 'POST':
        search_str = json.loads(request.body).get('searchText')
        
        try:
            date = parse_date(search_str)
        except ValueError:
            date = None

        if date:
            income = UserIncome.objects.filter(date=date, owner=request.user)
        else:
            income = UserIncome.objects.filter(
                amount__startswith=search_str, owner=request.user
            ) | UserIncome.objects.filter(
                description__icontains=search_str, owner=request.user
            ) | UserIncome.objects.filter(
                source__icontains=search_str, owner=request.user  
            )

        data = income.values('amount', 'source', 'description', 'date')

        return JsonResponse(list(data), safe=False)
    
    
    
@login_required(login_url='/authentication/login')
def income_source_summary(request):
    today_date = datetime.date.today()
    six = today_date - datetime.timedelta(days=180)
    income = UserIncome.objects.filter(owner=request.user, date__gte=six, date__lte=today_date)

    finalrep = {}

    def get_source(income):
        return income.source

    def get_income_source_amount(source):
        amount = 0
       
        filtered_by_source = income.filter(source=source)
        for inc in filtered_by_source:
            amount += inc.amount
        return amount

    source_list = list(set(map(get_source, income))) 
    for source in source_list:
        finalrep[source] = get_income_source_amount(source)

    return JsonResponse({'income_source_data': finalrep}, safe=False)




def stats_view(request):
    return render(request,'income/instats.html')


@login_required(login_url='/authentication/login')    
def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    date=str(datetime.datetime.now())
    response['Content-Disposition'] = 'attachment; filename=Income'+date+'.csv'
    writer=csv.writer(response)
    writer.writerow(['Amount','Source','Description','Date'])
    incomes = UserIncome.objects.filter(owner=request.user)
    for inc in incomes:
        writer.writerow([inc.amount, inc.source, inc.description, inc.date])
   
    return response


@login_required(login_url='/authentication/login')
def export_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date = str(datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
    response['Content-Disposition'] = f'attachment; filename=Income_{date}.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Expenses'
    ws['A1'] = 'Amount'
    ws['B1'] = 'Source'
    ws['C1'] = 'Description'
    ws['D1'] = 'Date'
    incomes = UserIncome.objects.filter(owner=request.user)
    
    for i, inc in enumerate(incomes, start=2):
        ws.cell(row=i, column=1).value = inc.amount
        ws.cell(row=i, column=2).value = inc.source
        ws.cell(row=i, column=3).value = inc.description
        
        date_cell = ws.cell(row=i, column=4)
        if isinstance(inc.date, datetime.date): 
            date_cell.value = inc.date
        else:
            
            date_cell.value = datetime.datetime.strptime(str(inc.date), "%Y-%m-%d")
        date_cell.number_format = 'YYYY-MM-DD'
        for col in range(1, 5): 
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        ws.column_dimensions[column].width = max_length + 2  
    wb.save(response)
    return response

@login_required(login_url='/authentication/login')
def export_pdf(request):
    incomes = UserIncome.objects.filter(owner=request.user)
    date = str(datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=Income_{date}.pdf'
    response['Content-Transfer-Encoding'] = "binary"
    incomes = UserIncome.objects.filter(owner=request.user)
    logo_url = request.build_absolute_uri(static('img/Triple_A_icon.jpg'))
    sum=incomes.aggregate(total=Sum('amount'))['total'] or 0
    html_string = render_to_string('income/pdf-output.html', {
        'incomes': incomes,
        'total': sum,
        'logo_url': logo_url
    })

    
    html = HTML(string=html_string)

    with tempfile.NamedTemporaryFile(delete=False) as output:
        html.write_pdf(output)
        output.seek(0)
        response.write(output.read())

    output.close()
    return response